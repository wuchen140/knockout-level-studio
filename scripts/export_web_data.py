#!/usr/bin/env python3
"""Convert the source level workbooks into lazily loaded web JSON files."""

from __future__ import annotations

import json
import math
import re
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
LEVEL_BOOK = Path("/Users/wuchen/Downloads/关卡配置.xlsx")
GAME_BOOK = Path("/Users/wuchen/Downloads/游戏配置.xlsx")
OUTPUT = ROOT / "public" / "data"

MATERIAL_NAMES = {
    0: "木头",
    1: "石头",
    2: "金属",
    3: "冰",
    4: "玻璃",
    5: "炸药",
    6: "塑料",
    7: "粉碎机",
    8: "糖果罐",
    9: "柱体",
}
SHAPE_NAMES = {0: "方块", 1: "圆柱", 2: "圆锥"}
COLOR_NAMES = {0: "无", 1: "红", 2: "黄", 3: "蓝", 4: "绿", 5: "橙", 6: "粉", 7: "紫"}


def clean(value):
    if isinstance(value, float):
        if math.isnan(value) or math.isinf(value):
            return 0
        return round(value, 6)
    return value


def slug(category: str, level_id: int) -> str:
    safe = re.sub(r"[^A-Za-z0-9_-]+", "-", str(category)).strip("-") or "level"
    return f"{safe}-{level_id}"


def rows_as_dicts(sheet):
    iterator = sheet.iter_rows(values_only=True)
    headers = next(iterator)
    for row in iterator:
        if not any(value not in (None, "") for value in row):
            continue
        yield {header: clean(value) for header, value in zip(headers, row)}


def vector(row, names):
    values = []
    for name in names:
        value = row.get(name)
        if value in (None, ""):
            values.append(0)
            continue
        try:
            values.append(float(value))
        except (TypeError, ValueError):
            values.append(0)
    return values


def quaternion_to_euler_degrees(q):
    x, y, z, w = q
    sinr_cosp = 2 * (w * x + y * z)
    cosr_cosp = 1 - 2 * (x * x + y * y)
    rx = math.atan2(sinr_cosp, cosr_cosp)
    sinp = 2 * (w * y - z * x)
    ry = math.copysign(math.pi / 2, sinp) if abs(sinp) >= 1 else math.asin(sinp)
    siny_cosp = 2 * (w * z + x * y)
    cosy_cosp = 1 - 2 * (y * y + z * z)
    rz = math.atan2(siny_cosp, cosy_cosp)
    return [round(math.degrees(value), 4) for value in (rx, ry, rz)]


def main():
    if not LEVEL_BOOK.exists() or not GAME_BOOK.exists():
        raise SystemExit("Missing /Users/wuchen/Downloads/关卡配置.xlsx or 游戏配置.xlsx")

    OUTPUT.mkdir(parents=True, exist_ok=True)
    levels_dir = OUTPUT / "levels"
    levels_dir.mkdir(parents=True, exist_ok=True)

    game = load_workbook(GAME_BOOK, read_only=True, data_only=True)
    profiles = []
    for index, row in enumerate(rows_as_dicts(game["方块档案"]), 1):
        profiles.append({
            "id": index,
            "material": row["材质"],
            "materialId": row["材质ID"],
            "shape": row["形状"],
            "shapeId": row["形状ID"],
            "size": vector(row, ["尺寸X", "尺寸Y", "尺寸Z"]),
            "mass": row["运行时质量"],
            "staticFriction": row["静摩擦"],
            "dynamicFriction": row["动摩擦"],
            "physicsMaterial": row["物理材质"],
            "impactShatter": row["撞击粉碎"] == "是",
            "shatterThreshold": row["粉碎速度阈值"],
        })

    colors = []
    seen_colors = set()
    for row in rows_as_dicts(game["颜色配置"]):
        key = (row["材质ID"], row["颜色ID"])
        if key in seen_colors:
            continue
        seen_colors.add(key)
        colors.append({
            "material": row["适用材质"],
            "materialId": row["材质ID"],
            "name": row["颜色"],
            "colorId": row["颜色ID"],
            "hex": row["基础色HEX"],
            "emission": row["自发光色HEX"],
            "highlight": row["高光色HEX"],
        })

    (OUTPUT / "catalog.json").write_text(
        json.dumps({"profiles": profiles, "colors": colors}, ensure_ascii=False, separators=(",", ":")),
        encoding="utf-8",
    )

    workbook = load_workbook(LEVEL_BOOK, read_only=True, data_only=True)
    summaries = {}
    order = []
    for row in rows_as_dicts(workbook["关卡配置"]):
        key = (str(row["关卡分类"]), int(row["关卡ID"]))
        item = {
            "key": f"{key[0]}:{key[1]}",
            "slug": slug(*key),
            "category": key[0],
            "id": key[1],
            "moveCount": row["移动次数"],
            "difficulty": row["难度名称"],
            "difficultyValue": row["难度值"],
            "progressionCount": row["进度引用次数"],
            "firstProgressionLevel": row["首次出现进度关卡"],
            "ballCount": row["球数"],
            "counts": {
                "platforms": row["平台数"],
                "blocks": row["基础方块数"],
                "barriers": row["障碍数"],
                "stages": row["阶段数"],
                "shutters": row["闸门数"],
                "waves": row["生成波次数"],
                "generatedBlocks": row["生成方块数"],
                "shutterBlocks": row["闸门方块数"],
            },
        }
        summaries[key] = item
        order.append(key)

    platforms = defaultdict(list)
    for row in rows_as_dicts(workbook["关卡平台明细"]):
        key = (str(row["关卡分类"]), int(row["关卡ID"]))
        q = vector(row, ["旋转X", "旋转Y", "旋转Z", "旋转W"])
        platforms[key].append({
            "uid": f"platform-{row['物品路径']}",
            "type": "platform",
            "name": f"平台 {row['平台序号']}",
            "area": row["所在区域"],
            "path": row["物品路径"],
            "stageIndex": row["阶段序号"],
            "platformIndex": row["平台序号"],
            "position": vector(row, ["位置X", "位置Y", "位置Z"]),
            "rotation": quaternion_to_euler_degrees(q),
            "size": vector(row, ["尺寸X", "尺寸Y", "尺寸Z"]),
            "motion": {
                "rotating": bool(row["是否旋转"]),
                "rotationSpeed": row["旋转速度"],
                "horizontal": bool(row["是否水平移动"]),
                "horizontalMin": row["水平最小位置"],
                "horizontalMax": row["水平最大位置"],
                "horizontalDirection": row["水平方向原名"],
                "horizontalSpeed": row["水平速度"],
                "vertical": bool(row["是否垂直移动"]),
                "verticalMin": row["垂直最小位置"],
                "verticalMax": row["垂直最大位置"],
                "verticalDirection": row["垂直方向原名"],
                "verticalSpeed": row["垂直速度"],
            },
        })

    blocks = defaultdict(list)
    for row in rows_as_dicts(workbook["关卡方块明细"]):
        key = (str(row["关卡分类"]), int(row["关卡ID"]))
        q = vector(row, ["旋转X", "旋转Y", "旋转Z", "旋转W"])
        material_id = int(row["材质值"] or 0)
        shape_id = int(row["形状值"] or 0)
        color_id = int(row["颜色值"] or 0)
        index = row["方块序号"]
        area = row["所在区域"]
        blocks[key].append({
            "uid": f"block-{area}-{row['阶段序号']}-{row['平台序号']}-{row['生成波次序号']}-{row['闸门序号']}-{index}",
            "type": "block",
            "name": f"方块 {index}",
            "area": area,
            "stageIndex": row["阶段序号"],
            "platformIndex": row["平台序号"],
            "waveIndex": row["生成波次序号"],
            "shutterIndex": row["闸门序号"],
            "blockIndex": index,
            "materialId": material_id,
            "materialName": MATERIAL_NAMES.get(material_id, f"材质 {material_id}"),
            "shapeId": shape_id,
            "shapeName": SHAPE_NAMES.get(shape_id, f"形状 {shape_id}"),
            "colorId": color_id,
            "colorName": COLOR_NAMES.get(color_id, row["颜色原名"] or f"颜色 {color_id}"),
            "position": vector(row, ["位置X", "位置Y", "位置Z"]),
            "rotation": quaternion_to_euler_degrees(q),
            "size": vector(row, ["尺寸X", "尺寸Y", "尺寸Z"]),
        })

    index = []
    for key in order:
        summary = summaries[key]
        data = {
            "schemaVersion": 1,
            "source": {"levels": LEVEL_BOOK.name, "game": GAME_BOOK.name},
            **summary,
            "objects": platforms[key] + blocks[key],
        }
        (levels_dir / f"{summary['slug']}.json").write_text(
            json.dumps(data, ensure_ascii=False, separators=(",", ":")), encoding="utf-8"
        )
        index.append(summary)

    (OUTPUT / "index.json").write_text(
        json.dumps({"levels": index}, ensure_ascii=False, separators=(",", ":")), encoding="utf-8"
    )
    print(f"Exported {len(index)} levels, {sum(map(len, blocks.values()))} blocks, {sum(map(len, platforms.values()))} platforms")


if __name__ == "__main__":
    main()
